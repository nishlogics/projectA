"""
pfd_scaleup.py
==============
Scale-up + mass-balance engine and a CTGR3-style workbook writer.

Design decisions (confirmed with the user):
  * Every raw material's plant quantity = lab quantity x scale-up ratio,
    UNLESS the material is flagged manual (stoichiometric / solution make-up),
    in which case the user supplies the plant quantity directly.
  * Scale-up ratio is entered manually (not derived).
  * Reference moles (for any equivalents the user computes by hand) are manual.
  * Output layout mirrors CTGR3.xlsx: top calc block + raw-material table
    (lab gm/ml | plant Kg/L | MW | moles | w/w | v/w | density | remarks)
    followed by the flowchart region.

The top-block mass balance (as reverse-engineered from CTGR3.xlsx):
    scale_up_ratio      : manual
    plant_batch_size    : lab_input * ratio   (or entered)
    input_moles         : plant_batch_size / input_MW      [manual override allowed]
    output_mass         : input_moles * yield * product_MW
    output_mass  -> becomes the next stage's plant_batch_size (chained).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from pfd_pipeline import Step, is_decision


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
@dataclass
class RawMaterial:
    sr_no: int
    name: str
    lab_qty_g: Optional[float] = None    # solid lab charge (gm)
    lab_vol_ml: Optional[float] = None   # liquid lab charge (ml)
    mw: Optional[float] = None
    density: Optional[float] = None
    remarks: str = ""
    manual: bool = False                 # True -> stoichiometric/solution: user gives plant qty
    manual_plant_kg: Optional[float] = None
    manual_plant_l: Optional[float] = None

    # computed
    plant_kg: Optional[float] = None
    plant_l: Optional[float] = None
    moles: Optional[float] = None
    w_w: Optional[float] = None          # weight/weight vs batch size
    v_w: Optional[float] = None          # volume/weight vs batch size

    def scale(self, ratio: float, batch_size_kg: float) -> None:
        """Apply plant quantities. Simple-scaled unless flagged manual."""
        if self.manual:
            self.plant_kg = self.manual_plant_kg
            self.plant_l = self.manual_plant_l
        else:
            self.plant_kg = (self.lab_qty_g * ratio / 1000.0) if self.lab_qty_g else None
            self.plant_l = (self.lab_vol_ml * ratio / 1000.0) if self.lab_vol_ml else None

        # moles from solid mass if MW known
        if self.plant_kg and self.mw:
            self.moles = self.plant_kg / self.mw
        # ratios vs batch size
        if batch_size_kg:
            if self.plant_kg is not None:
                self.w_w = self.plant_kg / batch_size_kg
            if self.plant_l is not None:
                self.v_w = self.plant_l / batch_size_kg


@dataclass
class StageCalc:
    name: str
    lab_input_g: float           # lab charge of the starting material (gm)
    input_mw: float
    product_mw: float
    scale_up_ratio: float        # manual
    yield_frac: float            # manual (e.g. 0.8)
    project: str = ""
    plant_batch_kg: Optional[float] = None   # if None -> lab_input_g*ratio/1000
    input_moles_override: Optional[float] = None
    materials: list[RawMaterial] = field(default_factory=list)
    steps: list[Step] = field(default_factory=list)

    # computed
    input_moles: Optional[float] = None
    output_kg: Optional[float] = None
    output_moles: Optional[float] = None

    def compute(self) -> None:
        if self.plant_batch_kg is None:
            self.plant_batch_kg = self.lab_input_g * self.scale_up_ratio / 1000.0
        self.input_moles = (self.input_moles_override
                            if self.input_moles_override is not None
                            else self.plant_batch_kg / self.input_mw)
        self.output_moles = self.input_moles * self.yield_frac
        self.output_kg = self.output_moles * self.product_mw
        for m in self.materials:
            m.scale(self.scale_up_ratio, self.plant_batch_kg)


def chain_stages(stages: list[StageCalc]) -> None:
    """Compute stages in order; each stage's output feeds the next stage's
    plant batch size (and lab input stays as entered for the ratio math)."""
    prev_out = None
    for s in stages:
        if prev_out is not None and s.plant_batch_kg is None:
            s.plant_batch_kg = prev_out
        s.compute()
        prev_out = s.output_kg


# ---------------------------------------------------------------------------
# CTGR3-style workbook writer
# ---------------------------------------------------------------------------
_GREEN = PatternFill("solid", fgColor="92D050")
_GREY = PatternFill("solid", fgColor="F2F2F2")
_YELLOW = PatternFill("solid", fgColor="FFFF00")
_BLUE = PatternFill("solid", fgColor="0070C0")
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_F = "Times New Roman"


def _cell(ws, coord, value, *, bold=False, fill=None, border=False,
          align=_CENTER, size=12, color="000000"):
    c = ws[coord]
    c.value = value
    c.font = Font(name=_F, bold=bold, size=size, color=color)
    c.alignment = align
    if fill:
        c.fill = fill
    if border:
        c.border = _BORDER
    return c


def write_stage_sheet(ws, stage: StageCalc) -> None:
    # column widths to match reference
    widths = {"A": 24.5, "B": 31, "C": 18.5, "D": 18.7, "E": 14.9,
              "F": 17, "G": 14.9, "H": 12, "I": 15.4, "J": 15.7, "K": 12.3, "L": 17}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---- header ----
    _cell(ws, "A4", "ISMS Class: INTERNAL", bold=True, align=_LEFT)
    _cell(ws, "L4", "ISMS Version: 1.0", bold=True, align=_LEFT)
    ws.merge_cells("A5:L5")
    _cell(ws, "A5", "Process Flow Diagram", bold=True, size=14)
    ws.merge_cells("A6:B6"); _cell(ws, "A6", "Project:", bold=True, align=_LEFT)
    _cell(ws, "C6", stage.project, align=_LEFT)
    _cell(ws, "A7", "Stage code:", bold=True, align=_LEFT)

    # ---- top calc block ----
    ws.merge_cells("A9:B9"); _cell(ws, "A9", "Batch Size", bold=True, align=_LEFT)
    _cell(ws, "C9", round(stage.plant_batch_kg, 3)); _cell(ws, "D9", "Kg")
    _cell(ws, "F9", "Output", bold=True, align=_LEFT)
    _cell(ws, "I9", round(stage.output_kg, 3)); _cell(ws, "J9", "Kg")

    _cell(ws, "A10", "Mol. Wt", bold=True, align=_LEFT)
    _cell(ws, "C10", stage.input_mw)
    _cell(ws, "F10", "Mol. Wt", bold=True, align=_LEFT)
    _cell(ws, "I10", stage.product_mw)

    _cell(ws, "A11", "Moles", bold=True, align=_LEFT)
    _cell(ws, "C11", round(stage.input_moles, 5))
    _cell(ws, "F11", "Moles", bold=True, align=_LEFT)
    _cell(ws, "I11", round(stage.output_moles, 5))

    _cell(ws, "A12", "Yield", bold=True, align=_LEFT)
    _cell(ws, "B12", "(50±20 %)", align=_LEFT)
    _cell(ws, "C12", stage.yield_frac)
    _cell(ws, "F12", "Scale Up ratio", bold=True, align=_LEFT)
    _cell(ws, "I12", stage.scale_up_ratio)

    # ---- raw material table ----
    ws.merge_cells("A14:A15"); _cell(ws, "A14", "Sr. No.", bold=True, fill=_GREEN, border=True)
    ws.merge_cells("B14:B15"); _cell(ws, "B14", "Raw material", bold=True, fill=_GREEN, border=True)
    ws.merge_cells("C14:D14"); _cell(ws, "C14", "Lab Batch", bold=True, fill=_GREEN, border=True)
    ws.merge_cells("E14:F14"); _cell(ws, "E14", "Plant Batch", bold=True, fill=_GREEN, border=True)
    for coord, txt in [("G14", "Mol. Wt.\nKg/Kgmol"), ("H14", "Moles"),
                       ("I14", "W/w"), ("J14", "V/W"), ("K14", "Density"), ("L14", "Remarks")]:
        ws.merge_cells(f"{coord[0]}14:{coord[0]}15")
        _cell(ws, coord, txt, bold=True, fill=_GREEN, border=True)
    _cell(ws, "C15", "Qty. (gm)", bold=True, fill=_GREEN, border=True)
    _cell(ws, "D15", "Vol. (ml)", bold=True, fill=_GREEN, border=True)
    _cell(ws, "E15", "Qty. (Kg)", bold=True, fill=_GREEN, border=True)
    _cell(ws, "F15", "Vol. (L)", bold=True, fill=_GREEN, border=True)

    r = 16
    for m in stage.materials:
        _cell(ws, f"A{r}", m.sr_no, border=True)
        _cell(ws, f"B{r}", m.name, border=True, align=_LEFT)
        _cell(ws, f"C{r}", _round(m.lab_qty_g), border=True)
        _cell(ws, f"D{r}", _round(m.lab_vol_ml), border=True)
        _cell(ws, f"E{r}", _round(m.plant_kg), border=True)
        _cell(ws, f"F{r}", _round(m.plant_l), border=True)
        _cell(ws, f"G{r}", _round(m.mw), border=True)
        _cell(ws, f"H{r}", _round(m.moles, 4), border=True)
        _cell(ws, f"I{r}", _round(m.w_w, 4), border=True)
        _cell(ws, f"J{r}", _round(m.v_w, 3), border=True)
        _cell(ws, f"K{r}", _round(m.density), border=True)
        _cell(ws, f"L{r}", m.remarks, border=True, align=_LEFT)
        r += 1

    # ---- flowchart region (below the table) ----
    _write_flowchart(ws, stage.steps, start_row=r + 2)


def _write_flowchart(ws, steps: list[Step], start_row: int) -> None:
    """Render a clean, professional vertical PFD.

    Layout per operation (3 worksheet rows tall):
        col A-B : material(s) charged + quantity  (input annotation, left)
        col D-F : operation box  (grey rectangle)  OR  blue decision diamond
        col G   : condition text (yellow) for decision steps
        col J-K : running Min / Max volume (right annotation)
        a centred down-arrow sits on the spacer row between consecutive boxes,
        and a short connector '─►' links the left material annotation into the box.
    """
    # section title
    _cell(ws, f"A{start_row}", "PROCESS FLOW DIAGRAM", bold=True, size=13, align=_LEFT)
    row = start_row + 2

    # header strip for the annotation columns
    _cell(ws, f"A{row}", "Material charged", bold=True, fill=_GREEN, border=True, size=10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
    _cell(ws, f"B{row}", "Qty", bold=True, fill=_GREEN, border=True, size=10)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    _cell(ws, f"D{row}", "Operation", bold=True, fill=_GREEN, border=True, size=10)
    _cell(ws, f"G{row}", "Decision", bold=True, fill=_GREEN, border=True, size=10)
    _cell(ws, f"J{row}", "Min Vol (L)", bold=True, fill=_GREEN, border=True, size=10)
    _cell(ws, f"K{row}", "Max Vol (L)", bold=True, fill=_GREEN, border=True, size=10)
    row += 2

    thick = Side(style="medium", color="000000")
    box_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    n = len(steps)
    for idx, s in enumerate(steps):
        ws.row_dimensions[row].height = 42
        ws.row_dimensions[row + 1].height = 6

        # ---- operation box (merged D:F across two rows) ----
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        box = ws.cell(row=row, column=4, value=f"{s.number}.  {s.full_text}")
        box.alignment = _CENTER

        if is_decision(s):
            box.font = Font(name=_F, bold=True, size=11, color="FFFFFF")
            box.fill = _BLUE
            box.border = box_border
            # decision text in yellow to the right
            cd = ws.cell(row=row, column=7, value=s.condition)
            cd.fill = _YELLOW
            cd.border = _BORDER
            cd.alignment = _LEFT
            cd.font = Font(name=_F, size=10, italic=True)
        else:
            box.font = Font(name=_F, bold=True, size=11)
            box.fill = _GREY
            box.border = box_border

        # ---- left annotation: materials + qty feeding this step ----
        if s.materials:
            mat = ws.cell(row=row, column=1, value=", ".join(s.materials))
            mat.alignment = _LEFT
            mat.font = Font(name=_F, size=10)
            mat.border = _BORDER
            qty = ", ".join(f"{q:g}" for q in s.quantities if q)
            if qty:
                qcell = ws.cell(row=row, column=2, value=qty)
                qcell.alignment = _CENTER
                qcell.font = Font(name=_F, size=10)
                qcell.border = _BORDER
            # connector from material into the box
            conn = ws.cell(row=row, column=3, value="──►")
            conn.alignment = Alignment(horizontal="right", vertical="center")
            conn.font = Font(name=_F, size=11, bold=True)

        # ---- right annotation: running volume ----
        if s.min_vol is not None:
            mv = ws.cell(row=row, column=10, value=round(s.min_vol, 2))
            mv.alignment = _CENTER; mv.border = _BORDER
            mv.font = Font(name=_F, size=10)
        if s.max_vol is not None:
            xv = ws.cell(row=row, column=11, value=round(s.max_vol, 2))
            xv.alignment = _CENTER; xv.border = _BORDER
            xv.font = Font(name=_F, size=10)

        # ---- down arrow on the spacer row (col E, centred under the box) ----
        if idx < n - 1:
            arr = ws.cell(row=row + 2, column=5, value="▼")
            arr.alignment = _CENTER
            arr.font = Font(name=_F, size=14, bold=True, color="0070C0")
            ws.row_dimensions[row + 2].height = 18

        row += 3

    # freeze the header block so long flowcharts stay readable
    ws.sheet_view.showGridLines = False


def _round(v, n=2):
    if v is None:
        return None
    return round(v, n)


def build_workbook(stages: list[StageCalc], path_or_buffer) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for i, stage in enumerate(stages, start=1):
        ws = wb.create_sheet(title=stage.name or f"Stage{i}")
        write_stage_sheet(ws, stage)
    wb.save(path_or_buffer)
