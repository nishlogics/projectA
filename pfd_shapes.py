"""
pfd_shapes.py
=============
Build a PFD workbook with TRUE auto-shapes (rectangles for operations, diamonds
for decisions) joined by down-arrows — by writing DrawingML directly and injecting
it into the xlsx produced by openpyxl.

This renders identically in Excel and LibreOffice and matches the reference PFD's
shape style. General-purpose; no project-specific names.
"""

from __future__ import annotations

import io
import re
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from pfd_scaleup import StageCalc
from pfd_pipeline import is_decision

EMU = 9525  # per pixel

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR = PatternFill("solid", fgColor="E7E6E6")
_C = Alignment("center", "center", wrap_text=True)
_L = Alignment("left", "center", wrap_text=True)


def _fmt(v, n=2):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return f"{v:g}" if float(v) == int(v) else f"{round(v, n):g}"
    return str(v)


# ---------------------------------------------------------------------------
def build(stages: list[StageCalc]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    # collect shapes per sheet: list of (kind, x, y, w, h, text, fill, fontcolor)
    sheet_shapes: dict[str, list] = {}

    for i, stage in enumerate(stages, 1):
        title = (stage.name or f"Stage {i}")[:31]
        ws = wb.create_sheet(title=title)
        shapes = _write_sheet(ws, stage)
        sheet_shapes[title] = shapes

    # save then inject drawings
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _inject_drawings(buf.getvalue(), wb, sheet_shapes)


def _write_sheet(ws, stage: StageCalc) -> list:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 11

    b = Font(bold=True)
    ws["B1"] = "Process Flow Diagram"; ws["B1"].font = Font(bold=True, size=13)

    ws["B3"] = "Batch size"; ws["B3"].font = b; ws["C3"] = _fmt(stage.plant_batch_kg); ws["D3"] = "Kg"
    ws["F3"] = "Output"; ws["F3"].font = b; ws["G3"] = _fmt(stage.output_kg); ws["H3"] = "Kg"
    ws["B4"] = "Input MW"; ws["B4"].font = b; ws["C4"] = _fmt(stage.input_mw)
    ws["F4"] = "Product MW"; ws["F4"].font = b; ws["G4"] = _fmt(stage.product_mw)
    ws["B5"] = "Moles"; ws["B5"].font = b; ws["C5"] = _fmt(stage.moles_report, 4)
    ws["F5"] = "Yield"; ws["F5"].font = b; ws["G5"] = _fmt(stage.yield_frac)

    # raw material table
    tr = 7
    for c, h in enumerate(["S.No", "Raw material", "Qty", "Unit", "MW",
                           "Mole ratio", "Plant Kg", "Plant L"], 1):
        cell = ws.cell(row=tr, column=c, value=h)
        cell.font = b; cell.fill = _HDR; cell.border = _BORDER; cell.alignment = _C
    tr += 1
    for m in stage.materials:
        qty = m.lab_qty_g if m.lab_qty_g else m.lab_vol_ml
        unit = "g" if m.lab_qty_g else ("mL" if m.lab_vol_ml else "")
        vals = [m.sr_no, m.name, _fmt(qty), unit, _fmt(m.mw),
                m.remarks, _fmt(m.plant_kg), _fmt(m.plant_l)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=tr, column=c, value=v)
            cell.border = _BORDER
            cell.alignment = _L if c == 2 else _C
        tr += 1

    # --- compute shape geometry (pixels) ---
    shapes = []
    # start flow below the table
    top = (tr + 2) * 20            # approx row height 20px
    x = 250
    box_w, box_h, dia_h, gap = 300, 52, 74, 26
    y = top

    steps = stage.steps
    for idx, s in enumerate(steps):
        decision = is_decision(s)
        h = dia_h if decision else box_h
        kind = "diamond" if decision else "rect"
        fill = "0070C0" if decision else "F2F2F2"
        fontc = "FFFFFF" if decision else "000000"
        shapes.append((kind, x, y, box_w, h, s.full_text, fill, fontc))

        # material label (left, no border) as a textbox rect with no fill
        if s.materials:
            q = ", ".join(_fmt(v) for v in s.quantities if v)
            lbl = ", ".join(s.materials) + (f"  ({q})" if q else "")
            shapes.append(("label", 10, y, 235, h, lbl, None, "444444"))

        # volume ledger (right)
        if s.max_vol is not None:
            shapes.append(("label", x + box_w + 12, y, 95, h,
                           f"{_fmt(s.min_vol)}-{_fmt(s.max_vol)} L", None, "444444"))

        # arrow down
        if idx < len(steps) - 1:
            shapes.append(("arrow", x + box_w // 2 - 8, y + h, 16, gap, "", "000000", None))
        y += h + gap

    return shapes


# ---------------------------------------------------------------------------
def _sp_xml(sid, kind, x, y, w, h, text, fill, fontc):
    ox, oy, cx, cy = x * EMU, y * EMU, w * EMU, h * EMU

    if kind == "arrow":
        return f"""<xdr:sp macro="" textlink=""><xdr:nvSpPr>
<xdr:cNvPr id="{sid}" name="Arrow {sid}"/><xdr:cNvSpPr/></xdr:nvSpPr>
<xdr:spPr><a:xfrm><a:off x="{ox}" y="{oy}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>
<a:prstGeom prst="downArrow"><a:avLst/></a:prstGeom>
<a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>
<a:ln><a:solidFill><a:srgbClr val="{fill}"/></a:solidFill></a:ln></xdr:spPr>
<xdr:txBody><a:bodyPr/><a:p><a:endParaRPr lang="en-US"/></a:p></xdr:txBody></xdr:sp>"""

    if kind == "label":
        return f"""<xdr:sp macro="" textlink=""><xdr:nvSpPr>
<xdr:cNvPr id="{sid}" name="Label {sid}"/><xdr:cNvSpPr txBox="1"/></xdr:nvSpPr>
<xdr:spPr><a:xfrm><a:off x="{ox}" y="{oy}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>
<a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>
<xdr:txBody><a:bodyPr wrap="square" rtlCol="0" anchor="ctr"/>
<a:p><a:pPr algn="r"/><a:r><a:rPr lang="en-US" sz="900"><a:solidFill>
<a:srgbClr val="{fontc}"/></a:solidFill></a:rPr><a:t>{_esc(text)}</a:t></a:r></a:p></xdr:txBody></xdr:sp>"""

    # rect or diamond box
    return f"""<xdr:sp macro="" textlink=""><xdr:nvSpPr>
<xdr:cNvPr id="{sid}" name="Box {sid}"/><xdr:cNvSpPr/></xdr:nvSpPr>
<xdr:spPr><a:xfrm><a:off x="{ox}" y="{oy}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>
<a:prstGeom prst="{kind}"><a:avLst/></a:prstGeom>
<a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>
<a:ln w="12700"><a:solidFill><a:srgbClr val="000000"/></a:solidFill></a:ln></xdr:spPr>
<xdr:txBody><a:bodyPr wrap="square" rtlCol="0" anchor="ctr"/>
<a:p><a:pPr algn="ctr"/><a:r><a:rPr lang="en-US" sz="1000" b="1"><a:solidFill>
<a:srgbClr val="{fontc}"/></a:solidFill></a:rPr><a:t>{_esc(text)}</a:t></a:r></a:p></xdr:txBody></xdr:sp>"""


def _esc(t):
    return (str(t).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _drawing_xml(shapes):
    anchors = []
    sid = 2
    for (kind, x, y, w, h, text, fill, fontc) in shapes:
        ox, oy, cx, cy = x * EMU, y * EMU, w * EMU, h * EMU
        sp = _sp_xml(sid, kind, x, y, w, h, text, fill, fontc)
        anchors.append(
            f'<xdr:absoluteAnchor><xdr:pos x="{ox}" y="{oy}"/>'
            f'<xdr:ext cx="{cx}" cy="{cy}"/>{sp}<xdr:clientData/></xdr:absoluteAnchor>')
        sid += 1
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
            'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
            + "".join(anchors) + "</xdr:wsDr>")


def _inject_drawings(xlsx_bytes, wb, sheet_shapes) -> bytes:
    """Post-process the xlsx zip: add drawingN.xml per sheet and wire the rels."""
    zin = zipfile.ZipFile(io.BytesIO(xlsx_bytes))

    # map sheet title -> sheetN.xml (order of wb.worksheets matches sheet1,2,..)
    sheet_files = {}
    for idx, ws in enumerate(wb.worksheets, 1):
        sheet_files[ws.title] = f"xl/worksheets/sheet{idx}.xml"

    # assign a drawing file to each non-empty sheet
    drawings = {}   # sheet_xml_path -> (drawing_path, drawing_xml, idx)
    didx = 0
    for title, shapes in sheet_shapes.items():
        if not shapes:
            continue
        didx += 1
        drawings[sheet_files[title]] = (
            f"xl/drawings/drawing{didx}.xml", _drawing_xml(shapes), didx)

    out = io.BytesIO()
    zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)

    for item in zin.infolist():
        data = zin.read(item.filename)

        # 1) content types: add drawing override
        if item.filename == "[Content_Types].xml":
            txt = data.decode("utf-8")
            if drawings and "drawing+xml" not in txt:
                ov = "".join(
                    f'<Override PartName="/{d[0]}" '
                    'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    for d in drawings.values())
                txt = txt.replace("</Types>", ov + "</Types>")
            zout.writestr(item, txt.encode("utf-8"))
            continue

        # 2) sheet xml: add <drawing r:id=.../> and declare the r: namespace
        if item.filename in drawings:
            dpath, dxml, dnum = drawings[item.filename]
            txt = data.decode("utf-8")
            if "xmlns:r=" not in txt:
                txt = txt.replace(
                    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
                    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                    'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"',
                    1)
            if "<drawing " not in txt:
                txt = txt.replace("</worksheet>", '<drawing r:id="rIdDraw1"/></worksheet>')
            zout.writestr(item, txt.encode("utf-8"))

            # sheet rels
            base = item.filename.split("/")[-1]
            rels_path = f"xl/worksheets/_rels/{base}.rels"
            rel_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                       '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                       f'<Relationship Id="rIdDraw1" '
                       'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                       f'Target="../drawings/drawing{dnum}.xml"/></Relationships>')
            zout.writestr(rels_path, rel_xml)
            zout.writestr(dpath, dxml)
            continue

        zout.writestr(item, data)

    zout.close()
    return out.getvalue()
