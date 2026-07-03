"""
pfd_pipeline.py
================
Pure-Python port of the original 5 MATLAB scripts (Step_1.m .. Step_5.m) plus the
Excel-VBA visual formatter (Format_All / FormatStepsWithConditionalRhombus /
Format_NonEmpty_Cells_In_ColumnD / FormatNumericCellsWithBox).

The pipeline converts a PR&D process-optimization report into a formatted
Process-Flow-Diagram (PFD) style Excel workbook, with zero MATLAB / VBA dependency.

Stage mapping (MATLAB -> Python):
    Step_1.m  ->  extract_steps()            : report text/docx -> list of (step, notes)
    Step_2.m  ->  number_steps()             : add step numbers + alternate-row layout
    Step_3.m  ->  extract_conditions()       : detect "If ..." conditional statements
    Step_4.m  ->  assign_materials()         : keyword-aware material + qty assignment
    Step_5.m  ->  compute_volumes()          : cumulative Min_Vol / Max_Vol ledger
    VBA       ->  write_formatted_workbook() : boxes, blue rhombus decisions, arrows

All the domain rules from the original scripts are preserved:
  * Notes belong to the previous step and are joined with " | ".
  * Operation N lands on Excel row (2*N)  -> alternate rows blank (visual gaps).
  * A step is a "decision/condition" step iff an "If ..." phrase is detected.
  * A material is placed at the first step where 'add' or 'charge' precedes its
    name (word-boundary match); fallback = first step it appears in at all.
  * Quantity: prefer solid Kg column, else liquid L column.
  * Min_Vol = previous step's Max_Vol, unless the step text contains "separate"
    (a reset point where the user supplies Min_Vol). Max_Vol = Min_Vol + Σqty.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
@dataclass
class Step:
    number: int
    text: str                       # instruction text (Column C)
    notes: list[str] = field(default_factory=list)
    condition: str = ""             # Column D  (extracted "If ..." phrases)
    materials: list[str] = field(default_factory=list)   # Column F
    quantities: list[float] = field(default_factory=list)  # Column G (aligned to materials)
    min_vol: Optional[float] = None  # Column H
    max_vol: Optional[float] = None  # Column I

    @property
    def full_text(self) -> str:
        """Instruction + inline notes joined with ' | ' (Step_1.m behaviour)."""
        if self.notes:
            return self.text + " | " + " | ".join(self.notes)
        return self.text

    @property
    def qty_sum(self) -> float:
        return float(sum(q for q in self.quantities if q is not None))


# ---------------------------------------------------------------------------
# STEP 1  (Step_1.m) : Report -> steps, with notes attached to previous step
# ---------------------------------------------------------------------------
_NUMBERED_RE = re.compile(r"^\s*(\d{1,3})[\.\)]\s+(.*)$")

# Lines that look like section headings, not process steps — skipped during parsing.
_HEADING_RE = re.compile(
    r"(?i)\b(process description|reaction procedure|synthetic scheme|"
    r"list of\s+raw\s+materials|list of raw materials|raw materials?:|scheme|"
    r"process\s+description:|draft process|theoretical yield|reported yield|"
    r"m\.?\s*wt|mole ratio)\b")


def _clean_line(line: str) -> str:
    return line.strip()


def extract_steps_from_lines(lines: list[str]) -> list[Step]:
    """
    Port of Step_1.m's state machine, generalised to also understand numbered
    lists (the report uses "1. ... 2. ..." rather than raw bullets).

    Rules:
      * A line starting with 'Note' is appended to the current step's notes.
      * A numbered line "N. text" starts a new step.
      * A plain non-empty, non-note line that is *not* numbered is treated as a
        continuation of the current step's instruction (wrapped report text).
    """
    steps: list[Step] = []
    current: Optional[Step] = None
    auto_no = 0

    for raw in lines:
        line = _clean_line(raw)
        if not line:
            continue

        if re.match(r"^note", line, re.IGNORECASE):
            if current is not None:
                current.notes.append(line)
            continue

        m = _NUMBERED_RE.match(line)
        if m:
            body = m.group(2).strip()
            # skip heading-like numbered lines (e.g. "1. CGTR1 Process Description:")
            if _HEADING_RE.search(body) and len(body) < 60:
                continue
            # finalise previous, start new (numbered)
            if current is not None:
                steps.append(current)
            auto_no = int(m.group(1))
            current = Step(number=auto_no, text=body)
        else:
            # skip standalone heading lines
            if _HEADING_RE.search(line) and current is None:
                continue
            # continuation of current step, or brand-new unnumbered step
            if current is None:
                auto_no += 1
                current = Step(number=auto_no, text=line)
            else:
                current.text = (current.text + " " + line).strip()

    if current is not None:
        steps.append(current)

    # renumber sequentially so downstream layout is deterministic
    for i, s in enumerate(steps, start=1):
        s.number = i
    return steps


def extract_steps_from_text(text: str) -> list[Step]:
    return extract_steps_from_lines(text.splitlines())


def extract_steps_from_pdf(file) -> list[Step]:
    """Read a report PDF, isolate the section after 'Procedure'/'Process description'
    if present, and parse numbered steps. Mirrors the Streamlit PDF path."""
    import pdfplumber

    with pdfplumber.open(file) as pdf:
        full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    m = re.search(r"(?i)\b(procedure|process description|reaction procedure)\b", full_text)
    if m:
        full_text = full_text[m.end():]
    return extract_steps_from_text(full_text)


def extract_steps_from_docx(file) -> list[Step]:
    from docx import Document

    doc = Document(file)
    lines = [p.text for p in doc.paragraphs]
    return extract_steps_from_lines(lines)


# ---------------------------------------------------------------------------
# STEP 2  (Step_2.m) : numbering + alternate-row layout
# ---------------------------------------------------------------------------
# (Handled implicitly by Step.number and the writer, which places operation N on
#  Excel row 2*N. Kept as an explicit function for parity / testability.)
def numbered_layout(steps: list[Step]) -> list[tuple[int, int, str]]:
    """Return (excel_row, step_number, text) with operation N on row 2*N."""
    return [(2 * s.number, s.number, s.full_text) for s in steps]


# ---------------------------------------------------------------------------
# STEP 3  (Step_3.m) : conditional-statement extraction
# ---------------------------------------------------------------------------
_DEFAULT_CONDITION_KEYWORDS = [
    "if ipc", "if complies", "if does not comply", "if ipc result complies",
    "if the result", "if not", "if it complies",
]
_IF_PHRASE_RE = re.compile(r"If[^.]+", re.IGNORECASE)


def extract_conditions(steps: list[Step],
                        keywords: Optional[list[str]] = None) -> None:
    """Fill Step.condition in place (Step_3.m). Fast keyword filter, then regex
    extraction of every 'If ... ' phrase up to the next period."""
    kws = [k.lower() for k in (keywords or _DEFAULT_CONDITION_KEYWORDS)]
    for s in steps:
        hay = s.full_text.lower()
        if any(k in hay for k in kws):
            matches = _IF_PHRASE_RE.findall(s.full_text)
            if matches:
                s.condition = "; ".join(m.strip() for m in matches)


# ---------------------------------------------------------------------------
# STEP 4  (Step_4.m) : keyword-aware material + quantity assignment
# ---------------------------------------------------------------------------
_KW_ADD = re.compile(r"\badd\b", re.IGNORECASE)
_KW_CHARGE = re.compile(r"\bcharge\b", re.IGNORECASE)


@dataclass
class MaterialRow:
    name: str
    qty_kg: Optional[float] = None   # solid  (material-sheet col E / "Qty. (Kg)")
    vol_l: Optional[float] = None    # liquid (material-sheet col F / "Vol. (L)")

    @property
    def unified_qty(self) -> Optional[float]:
        if self.qty_kg is not None and self.qty_kg > 0:
            return self.qty_kg
        if self.vol_l is not None and self.vol_l > 0:
            return self.vol_l
        return None


def load_material_sheet(file_or_df,
                        name_col: int = 1,
                        kg_col: int = 4,
                        l_col: int = 5) -> list[MaterialRow]:
    """
    Load a material sheet into MaterialRow records.

    Column indices are 0-based and default to the original MATLAB layout:
        name -> column B (idx 1), Kg -> column E (idx 4), L -> column F (idx 5).
    Any header rows / blank names are skipped.
    """
    if isinstance(file_or_df, pd.DataFrame):
        df = file_or_df
    else:
        df = pd.read_excel(file_or_df, header=None)

    rows: list[MaterialRow] = []
    for _, r in df.iterrows():
        try:
            name = r.iloc[name_col]
        except IndexError:
            continue
        if pd.isna(name):
            continue
        name = str(name).strip()
        if not name or name.lower() in {"raw material", "raw materials", "s. no", "s.no"}:
            continue

        def _num(idx):
            try:
                v = r.iloc[idx]
            except IndexError:
                return None
            v = pd.to_numeric(v, errors="coerce")
            return None if pd.isna(v) else float(v)

        rows.append(MaterialRow(name=name, qty_kg=_num(kg_col), vol_l=_num(l_col)))
    return rows


def assign_materials(steps: list[Step], materials: list[MaterialRow]) -> None:
    """Port of Step_4.m. Each material is placed once:
       Pass 1 -> first step where 'add'/'charge' occurs *before* the material name.
       Pass 2 -> fallback: first step where the material name appears at all."""
    step_lc = [s.text.lower() for s in steps]

    for mat in materials:
        term = mat.name.lower()
        qty = mat.unified_qty
        assigned = False

        # Pass 1: keyword-before-material
        for i, low in enumerate(step_lc):
            mi = low.find(term)
            if mi < 0:
                continue
            add_before = any(m.start() < mi for m in _KW_ADD.finditer(low))
            chg_before = any(m.start() < mi for m in _KW_CHARGE.finditer(low))
            if add_before or chg_before:
                steps[i].materials.append(mat.name)
                steps[i].quantities.append(qty if qty else 0.0)
                assigned = True
                break

        # Pass 2: fallback first-occurrence
        if not assigned:
            for i, low in enumerate(step_lc):
                if term in low:
                    steps[i].materials.append(mat.name)
                    steps[i].quantities.append(qty if qty else 0.0)
                    break


# ---------------------------------------------------------------------------
# STEP 5  (Step_5.m) : cumulative Min_Vol / Max_Vol ledger
# ---------------------------------------------------------------------------
_SEPARATE_RE = re.compile(r"\bseparate\b", re.IGNORECASE)


def compute_volumes(steps: list[Step],
                    separate_overrides: Optional[dict[int, float]] = None,
                    default_separate_min: float = 0.0) -> None:
    """
    Port of Step_5.m.

    For each step:
        Min_Vol = previous step's Max_Vol
                  UNLESS the step text contains 'separate', in which case
                  Min_Vol is taken from `separate_overrides[step.number]`
                  (or `default_separate_min` if not supplied).
        Max_Vol = Min_Vol + Σ(quantities in that step).
    """
    overrides = separate_overrides or {}
    prev_max = 0.0
    for s in steps:
        if _SEPARATE_RE.search(s.text):
            s.min_vol = float(overrides.get(s.number, default_separate_min))
        else:
            s.min_vol = prev_max
        s.max_vol = s.min_vol + s.qty_sum
        prev_max = s.max_vol


def separate_steps(steps: list[Step]) -> list[Step]:
    """Steps that will prompt the user for a Min_Vol (contain 'separate')."""
    return [s for s in steps if _SEPARATE_RE.search(s.text)]


def is_decision(step: Step) -> bool:
    return bool(step.condition.strip())


# ---------------------------------------------------------------------------
# Convenience: full pipeline -> tidy DataFrame (mirrors output_with_vols.xlsx)
# ---------------------------------------------------------------------------
def steps_to_dataframe(steps: list[Step]) -> pd.DataFrame:
    return pd.DataFrame([{
        "Step": s.number,
        "Operation": s.full_text,
        "Condition": s.condition,
        "Material": ", ".join(s.materials),
        "Qty": ", ".join(f"{q:g}" for q in s.quantities if q),
        "Min_Vol": s.min_vol,
        "Max_Vol": s.max_vol,
    } for s in steps])


def run_pipeline(steps: list[Step],
                 materials: Optional[list[MaterialRow]] = None,
                 condition_keywords: Optional[list[str]] = None,
                 separate_overrides: Optional[dict[int, float]] = None) -> list[Step]:
    """Run steps 3->5 on already-extracted steps (step 1/2 happen at ingest)."""
    extract_conditions(steps, condition_keywords)
    if materials:
        assign_materials(steps, materials)
    compute_volumes(steps, separate_overrides)
    return steps


# ---------------------------------------------------------------------------
# VBA replacement : write a formatted PFD workbook
# ---------------------------------------------------------------------------
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_GREY = PatternFill("solid", fgColor="F2F2F2")
_YELLOW = PatternFill("solid", fgColor="FFFF00")
_BLUE = PatternFill("solid", fgColor="0070C0")
_WHITE = PatternFill("solid", fgColor="FFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_formatted_workbook(steps: list[Step], path_or_buffer) -> None:
    """
    Reproduce the VBA Format_All visual output using openpyxl + drawings.

    Layout (matches Step_2.m / the VBA expectations):
        Column A : step number
        Column C : operation text (grey box)  OR blue rhombus if it's a decision
        Column D : condition text (yellow box)
        Column F : material name(s)
        Column G : quantity(-ies)          -> boxed if numeric
        Column H : Min_Vol                  -> boxed if numeric
        Column I : Max_Vol                  -> boxed if numeric
        Operation N is written on row 2*N; alternate rows stay blank.
        A downward arrow (shape) connects consecutive operation boxes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Process Flow"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 40
    for col in ("F", "G", "H", "I"):
        ws.column_dimensions[col].width = 18

    # header row for the numeric side-columns
    hdr = ["", "", "Operation", "Condition", "", "Material", "Qty", "Min Vol", "Max Vol"]
    for c, val in enumerate(hdr, start=1):
        cell = ws.cell(row=1, column=c, value=val)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER

    last_row = 1
    for s in steps:
        r = 2 * s.number
        last_row = max(last_row, r)
        ws.row_dimensions[r].height = 60

        ws.cell(row=r, column=1, value=s.number).alignment = _CENTER

        op_cell = ws.cell(row=r, column=3, value=s.full_text)
        op_cell.alignment = _CENTER
        op_cell.font = Font(bold=True)

        if is_decision(s):
            # decision step: leave the C cell text but visually flag it blue.
            # (openpyxl cannot embed text inside an auto-shape reliably, so we
            #  colour the cell blue + white bold text to represent the rhombus,
            #  and additionally drop a diamond shape over it.)
            op_cell.fill = _BLUE
            op_cell.font = Font(bold=True, color="FFFFFF")
            _add_diamond(ws, r)
            # condition text (col D) -> yellow box
            d = ws.cell(row=r, column=4, value=s.condition)
            d.fill = _YELLOW
            d.border = _BORDER
            d.alignment = _CENTER
        else:
            op_cell.fill = _GREY
            op_cell.border = _BORDER

        # material / qty / vols
        if s.materials:
            ws.cell(row=r, column=6, value=", ".join(s.materials)).alignment = _CENTER
        qtys = [q for q in s.quantities if q]
        if qtys:
            g = ws.cell(row=r, column=7, value=", ".join(f"{q:g}" for q in qtys))
            g.border = _BORDER
            g.alignment = _CENTER
        if s.min_vol is not None:
            h = ws.cell(row=r, column=8, value=round(s.min_vol, 3))
            h.border = _BORDER
            h.alignment = _CENTER
        if s.max_vol is not None:
            i_ = ws.cell(row=r, column=9, value=round(s.max_vol, 3))
            i_.border = _BORDER
            i_.alignment = _CENTER

    # arrows between consecutive operations
    _add_arrows(ws, steps)

    wb.save(path_or_buffer)


def _add_diamond(ws, row: int) -> None:
    """Visual decision marker.

    openpyxl cannot embed text inside a free auto-shape reliably, so a decision
    step is represented by a solid blue fill + bold white text on the C cell
    (done by the caller). This hook is kept as an extension point and is a
    deliberate no-op to keep the workbook 100% valid across Excel/LibreOffice."""
    return


def _add_arrows(ws, steps: list[Step]) -> None:
    """Draw a simple down-arrow connector between consecutive operation rows
    using cell characters (robust across Excel/LibreOffice), placed on the
    blank spacer row between two operations, column C."""
    for a, b in zip(steps, steps[1:]):
        gap_row = 2 * a.number + 1
        cell = ws.cell(row=gap_row, column=3, value="↓")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=16, bold=True)


# ---------------------------------------------------------------------------
# Optional: richer flowchart via xlsxwriter (real merged boxes + arrows)
# ---------------------------------------------------------------------------
def write_flowchart_xlsxwriter(steps: list[Step]) -> io.BytesIO:
    """
    Alternative renderer using xlsxwriter, closer to the original Streamlit
    'create_excel_with_flowchart_only' but driven by the full pipeline
    (decisions from Step_3, materials/qty from Step_4, volumes from Step_5).
    Returns an in-memory BytesIO.
    """
    import xlsxwriter

    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out, {"in_memory": True})
    ws = wb.add_worksheet("Process Flow")

    box = wb.add_format({"border": 2, "align": "center", "valign": "vcenter",
                         "text_wrap": True, "bold": True, "bg_color": "#F2F2F2"})
    diamond = wb.add_format({"border": 2, "align": "center", "valign": "vcenter",
                             "text_wrap": True, "bold": True,
                             "font_color": "white", "bg_color": "#0070C0"})
    yellow = wb.add_format({"border": 1, "align": "center", "valign": "vcenter",
                            "text_wrap": True, "bg_color": "#FFFF00"})
    numbox = wb.add_format({"border": 1, "align": "center", "valign": "vcenter"})
    arrow = wb.add_format({"align": "center", "valign": "vcenter",
                           "font_size": 20, "bold": True})

    ws.set_column("A:A", 6)
    ws.set_column("C:E", 25)
    ws.set_column("F:I", 16)

    row = 0
    for idx, s in enumerate(steps):
        ws.set_row(row, 90)
        ws.write(row, 0, s.number, numbox)

        text = s.full_text
        if is_decision(s):
            ws.merge_range(row, 2, row + 1, 4, text, diamond)
            ws.write(row, 5, s.condition, yellow)
        else:
            ws.merge_range(row, 2, row + 1, 4, text, box)

        if s.materials:
            ws.write(row, 6, ", ".join(s.materials), numbox)
        qtys = [q for q in s.quantities if q]
        if qtys:
            ws.write(row, 7, ", ".join(f"{q:g}" for q in qtys), numbox)
        if s.min_vol is not None:
            ws.write(row, 8, round(s.min_vol, 3), numbox)

        if idx < len(steps) - 1:
            ws.write(row + 2, 3, "↓", arrow)
        row += 3

    wb.close()
    out.seek(0)
    return out
