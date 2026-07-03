"""
pfd_formula_writer.py
=====================
Writes a fully FORMULA-LINKED, CTGR3-matching PFD workbook.

Everything is driven by a single scale-up-ratio cell (I12). Change it in Excel and
the whole sheet recalculates, exactly like the reference CTGR3.xlsx:

  Raw-material table
    Plant Kg  (col E) = =C{row}*I$12/1000      (lab gm * ratio / 1000)
    Plant L   (col F) = =D{row}*I$12/1000       (lab mL * ratio / 1000)
    Moles     (col H) = =E{row}/G{row}          (plant kg / MW)
    W/w       (col I) = =E{row}/$C$9             (plant kg / batch size)
    V/W       (col J) = =F{row}/$C$9             (plant L  / batch size)

  Top block
    Batch (C9)   = =C16*I$12/1000               (SM lab gm * ratio / 1000)
    Moles (C11)  = =C9/C10                       (batch / input MW)
    Out moles(I11) = =C11*C12                    (moles * yield)
    Output (I9)  = =I11*I10                       (out moles * product MW)

  Flowchart (equipment blocks + operations)
    material name  (A) = =B{matrow}
    quantity       (B) = =F{matrow}  (liquid)  or  =E{matrow}  (solid)
    solution vol   (B) = =B{above}/K{matrow}    (solute kg / density)
    Min Vol        (J) = previous block/op Max Vol
    Max Vol        (K) = =J{r}+<sum of qty cells charged in this block>
    Equip occupancy(F/I) = =E/C  and  =H/C
    Observations column left blank for manual entry.
"""

from __future__ import annotations

from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pfd_scaleup import StageCalc, RawMaterial
from pfd_pipeline import Step, is_decision

_GREEN = PatternFill("solid", fgColor="92D050")
_GREY = PatternFill("solid", fgColor="F2F2F2")
_YELLOW = PatternFill("solid", fgColor="FFFF00")
_BLUE = PatternFill("solid", fgColor="0070C0")
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_C = Alignment("center", "center", wrap_text=True)
_L = Alignment("left", "center", wrap_text=True)
_F = "Times New Roman"

# raw-material table starts at this row (matches reference)
MAT_START = 16
RATIO_CELL = "I$12"      # scale-up ratio
BATCH_CELL = "$C$9"      # plant batch size


def _w(ws, coord, val, *, bold=False, fill=None, border=False, align=_C, size=12, color="000000"):
    c = ws[coord]
    c.value = val
    c.font = Font(name=_F, bold=bold, size=size, color=color)
    c.alignment = align
    if fill:
        c.fill = fill
    if border:
        c.border = _BORDER
    return c


def _link_step_materials(stage: StageCalc) -> None:
    """Populate each step's .materials by matching material names in the step text
    (keyword-aware, same rule as the pipeline). Idempotent."""
    import pandas as pd
    from pfd_pipeline import load_material_sheet, assign_materials
    for s in stage.steps:
        s.materials = []
        s.quantities = []
    rows = [[m.sr_no, m.name, None, m.plant_kg, m.plant_l] for m in stage.materials]
    if not rows:
        return
    mats = load_material_sheet(pd.DataFrame(rows), name_col=1, kg_col=3, l_col=4)
    assign_materials(stage.steps, mats)


def write_stage(ws, stage: StageCalc, ratio: float) -> None:
    widths = {"A": 24.5, "B": 31, "C": 18.5, "D": 30, "E": 14.9, "F": 17,
              "G": 14.9, "H": 12, "I": 15.4, "J": 15.7, "K": 12.3, "L": 14}
    for col, wd in widths.items():
        ws.column_dimensions[col].width = wd

    # ---------- header ----------
    ws.merge_cells("A5:L5")
    _w(ws, "A5", "Process Flow Diagram", bold=True, size=14)
    _w(ws, "A6", "Project:", bold=True, align=_L)
    _w(ws, "C6", stage.project, align=_L)

    # ---------- top calc block (all formulas) ----------
    _w(ws, "A9", "Batch Size", bold=True, align=_L)
    _w(ws, "C9", f"=C{MAT_START}*{RATIO_CELL}/1000"); _w(ws, "D9", "Kg")
    _w(ws, "F9", "Output", bold=True, align=_L)
    _w(ws, "I9", "=I11*I10"); _w(ws, "J9", "Kg")

    _w(ws, "A10", "Mol. Wt", bold=True, align=_L); _w(ws, "C10", stage.input_mw)
    _w(ws, "F10", "Mol. Wt", bold=True, align=_L); _w(ws, "I10", stage.product_mw)

    _w(ws, "A11", "Moles", bold=True, align=_L); _w(ws, "C11", "=C9/C10")
    _w(ws, "F11", "Moles", bold=True, align=_L); _w(ws, "I11", "=C11*C12")

    _w(ws, "A12", "Yield", bold=True, align=_L); _w(ws, "C12", stage.yield_frac)
    _w(ws, "F12", "Scale Up ratio", bold=True, align=_L); _w(ws, "I12", ratio)

    # ---------- raw-material table ----------
    ws.merge_cells("A14:A15"); _w(ws, "A14", "Sr. No.", bold=True, fill=_GREEN, border=True)
    ws.merge_cells("B14:B15"); _w(ws, "B14", "Raw material", bold=True, fill=_GREEN, border=True)
    ws.merge_cells("C14:D14"); _w(ws, "C14", "Lab Batch", bold=True, fill=_GREEN, border=True)
    ws.merge_cells("E14:F14"); _w(ws, "E14", "Plant Batch", bold=True, fill=_GREEN, border=True)
    for coord, txt in [("G14", "Mol. Wt.\nKg/Kgmol"), ("H14", "Moles"),
                       ("I14", "W/w"), ("J14", "V/W"), ("K14", "Density"), ("L14", "Remarks")]:
        ws.merge_cells(f"{coord[0]}14:{coord[0]}15")
        _w(ws, coord, txt, bold=True, fill=_GREEN, border=True)
    _w(ws, "C15", "Qty. (gm)", bold=True, fill=_GREEN, border=True)
    _w(ws, "D15", "Vol. (ml)", bold=True, fill=_GREEN, border=True)
    _w(ws, "E15", "Qty. (Kg)", bold=True, fill=_GREEN, border=True)
    _w(ws, "F15", "Vol. (L)", bold=True, fill=_GREEN, border=True)

    # map material name -> its table row (for flowchart references)
    mat_row: dict[str, int] = {}
    r = MAT_START
    for m in stage.materials:
        mat_row[m.name] = r
        _w(ws, f"A{r}", m.sr_no, border=True)
        _w(ws, f"B{r}", m.name, border=True, align=_L)
        _w(ws, f"C{r}", m.lab_qty_g, border=True)
        _w(ws, f"D{r}", m.lab_vol_ml, border=True)
        # plant qty formulas driven by ratio
        _w(ws, f"E{r}", f"=C{r}*{RATIO_CELL}/1000" if m.lab_qty_g else None, border=True)
        _w(ws, f"F{r}", f"=D{r}*{RATIO_CELL}/1000" if m.lab_vol_ml else None, border=True)
        _w(ws, f"G{r}", m.mw, border=True)
        _w(ws, f"H{r}", f"=E{r}/G{r}" if (m.lab_qty_g and m.mw) else None, border=True)
        _w(ws, f"I{r}", f"=E{r}/{BATCH_CELL}" if m.lab_qty_g else None, border=True)
        _w(ws, f"J{r}", f"=F{r}/{BATCH_CELL}" if m.lab_vol_ml else None, border=True)
        _w(ws, f"K{r}", m.density, border=True)
        _w(ws, f"L{r}", m.remarks, border=True, align=_L)
        r += 1

    # ---------- flowchart ----------
    _link_step_materials(stage)
    _write_flowchart(ws, stage, mat_row, start_row=r + 2)


def _write_flowchart(ws, stage: StageCalc, mat_row: dict[str, int], start_row: int) -> None:
    """Emit equipment blocks + operation rows with live formulas.

    Grouping strategy: one equipment block header, then all operations. (A single
    reactor R-1 is assumed unless the steps mention 'filter'/'dry' — a reasonable
    default; the engineer can split blocks in Excel.) Materials referenced in a
    step are pulled from the raw-material table by row; volumes carry forward.
    """
    row = start_row
    # equipment block header
    _w(ws, f"A{row}", "Equipment grouping", bold=True, fill=_GREEN, border=True)
    _w(ws, f"B{row}", "Equipment ID", bold=True, fill=_GREEN, border=True)
    _w(ws, f"C{row}", "Capacity", bold=True, fill=_GREEN, border=True)
    _w(ws, f"D{row}", "UOM", bold=True, fill=_GREEN, border=True)
    _w(ws, f"G{row}", "Observations", bold=True, fill=_GREEN, border=True)
    _w(ws, f"J{row}", "Min. Vol. L", bold=True, fill=_GREEN, border=True)
    _w(ws, f"K{row}", "Max. Vol. L", bold=True, fill=_GREEN, border=True)
    row += 1
    block_hdr = row
    _w(ws, f"A{row}", "R-1", bold=True, border=True)
    _w(ws, f"B{row}", "", border=True)       # equipment id (fill after selection)
    _w(ws, f"C{row}", "", border=True)       # capacity
    _w(ws, f"D{row}", "L", border=True)
    row += 1

    prev_max_cell = None
    first_op_minvol_cell = None
    last_op_maxvol_cell = None

    for idx, s in enumerate(stage.steps):
        op_row = row
        ws.row_dimensions[op_row].height = 48

        # operation text box (col D), blue diamond if decision
        target = ws.cell(row=op_row, column=4, value=s.full_text)
        target.alignment = _C
        if is_decision(s):
            target.fill = _BLUE
            target.font = Font(name=_F, bold=True, size=11, color="FFFFFF")
            target.border = _BORDER
            cond = ws.cell(row=op_row, column=7, value=s.condition)  # G
            cond.fill = _YELLOW; cond.border = _BORDER; cond.alignment = _C
        else:
            target.fill = _GREY
            target.font = Font(name=_F, bold=True, size=11)
            target.border = _BORDER

        # material + qty references (may be several materials in one step)
        qty_cells = []
        sub = op_row
        for mname in s.materials:
            mr = mat_row.get(mname)
            if mr is None:
                continue
            _w(ws, f"A{sub}", f"=B{mr}", border=True, align=_L)
            # liquid vs solid: if plant L exists use F, else solid E (+ solution row)
            m = next((x for x in stage.materials if x.name == mname), None)
            if m and m.lab_vol_ml:
                _w(ws, f"B{sub}", f"=F{mr}", border=True)
                _w(ws, f"C{sub}", "L", border=True)
                qty_cells.append(f"B{sub}")
            else:
                _w(ws, f"B{sub}", f"=E{mr}", border=True)
                _w(ws, f"C{sub}", "Kg", border=True)
                # if it has density, add a solution-volume sub-row (kg / density)
                if m and m.density:
                    sub += 1
                    _w(ws, f"B{sub}", f"=B{sub-1}/K{mr}", border=True)
                    _w(ws, f"C{sub}", "L", border=True)
                    qty_cells.append(f"B{sub}")
                else:
                    qty_cells.append(f"B{sub}")
            sub += 1

        # min / max vol ledger (live formulas)
        minvol = ws.cell(row=op_row, column=10)   # J
        maxvol = ws.cell(row=op_row, column=11)   # K
        minvol.alignment = _C; maxvol.alignment = _C
        if prev_max_cell is None:
            minvol.value = 0
        else:
            minvol.value = f"={prev_max_cell}"
        if qty_cells:
            maxvol.value = f"=J{op_row}+" + "+".join(qty_cells)
        else:
            maxvol.value = f"=J{op_row}"
        prev_max_cell = f"K{op_row}"
        if first_op_minvol_cell is None:
            first_op_minvol_cell = f"J{op_row}"
        last_op_maxvol_cell = f"K{op_row}"

        # arrow on the spacer row
        if idx < len(stage.steps) - 1:
            a = ws.cell(row=max(sub, op_row) + 1, column=4, value="↓")
            a.alignment = _C
            a.font = Font(name=_F, size=14, bold=True)

        row = max(sub, op_row) + 2

    # equipment occupancy formulas on the block header (min/max working vol / capacity)
    # guarded against a blank capacity cell so the sheet stays clean until an
    # equipment ID + capacity is filled in.
    if first_op_minvol_cell and last_op_maxvol_cell:
        cap = f"C{block_hdr}"
        _w(ws, f"E{block_hdr}", f"={first_op_minvol_cell}", border=True)
        _w(ws, f"F{block_hdr}", f'=IF({cap}="","",E{block_hdr}/{cap})', border=True)
        _w(ws, f"H{block_hdr}", f"={last_op_maxvol_cell}", border=True)
        _w(ws, f"I{block_hdr}", f'=IF({cap}="","",H{block_hdr}/{cap})', border=True)


def build_formula_workbook(stages: list[StageCalc], ratios: list[float], path_or_buffer) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for i, (stage, ratio) in enumerate(zip(stages, ratios), start=1):
        ws = wb.create_sheet(title=stage.name or f"Stage{i}")
        write_stage(ws, stage, ratio)
    wb.save(path_or_buffer)
